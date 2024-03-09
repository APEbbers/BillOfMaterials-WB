# ***************************************************************************
# *   Copyright (c) 2023 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/

import FreeCAD as App
import math

# Define the translation
translate = App.Qt.translate


def Mbox(text, title="", style=0, IconType="Information", default="", stringList="[,]"):
    """
    Message Styles:\n
    0 : OK                          (text, title, style)\n
    1 : Yes | No                    (text, title, style)\n
    2 : Ok | Cancel                 (text, title, style)\n
    20 : Inputbox                   (text, title, style, default)\n
    21 : Inputbox with dropdown     (text, title, style, default, stringlist)\n
    Icontype:                       string: NoIcon, Question, Warning, Critical. Default Information
    """
    from PySide.QtWidgets import QMessageBox, QInputDialog

    Icon = QMessageBox.Information
    if IconType == "NoIcon":
        Icon = QMessageBox.NoIcon
    if IconType == "Question":
        Icon = QMessageBox.Question
    if IconType == "Warning":
        Icon = QMessageBox.Warning
    if IconType == "Critical":
        Icon = QMessageBox.Critical

    if style == 0:
        # Set the messagebox
        msgBox = QMessageBox()
        msgBox.setIcon(Icon)
        msgBox.setText(text)
        msgBox.setWindowTitle(title)

        reply = msgBox.exec_()
        if reply == QMessageBox.Ok:
            return "ok"
    if style == 1:
        # Set the messagebox
        msgBox = QMessageBox()
        msgBox.setIcon(Icon)
        msgBox.setText(text)
        msgBox.setWindowTitle(title)
        # Set the buttons and default button
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msgBox.setDefaultButton(QMessageBox.No)

        reply = msgBox.exec_()
        if reply == QMessageBox.Yes:
            return "yes"
        if reply == QMessageBox.No:
            return "no"
    if style == 2:
        # Set the messagebox
        msgBox = QMessageBox()
        msgBox.setIcon(Icon)
        msgBox.setText(text)
        msgBox.setWindowTitle(title)
        # Set the buttons and default button
        msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        msgBox.setDefaultButton(QMessageBox.Ok)

        reply = msgBox.exec_()
        if reply == QMessageBox.Ok:
            return "ok"
        if reply == QMessageBox.Cancel:
            return "cancel"
    if style == 20:
        reply = QInputDialog.getText(
            None,
            title,
            text,
            text=default,
        )
        if reply[1]:
            # user clicked OK
            replyText = reply[0]
        else:
            # user clicked Cancel
            replyText = reply[0]  # which will be "" if they clicked Cancel
        return str(replyText)
    if style == 21:
        reply = QInputDialog.getItem(
            None,
            title,
            text,
            stringList,
            0,
            True,
        )
        if reply[1]:
            # user clicked OK
            replyText = reply[0]
        else:
            # user clicked Cancel
            replyText = reply[0]  # which will be "" if they clicked Cancel
        return str(replyText)


def SaveDialog(files, OverWrite: bool = True):
    """
    files must be like:\n
    files = [\n
        ('All Files', '*.*'),\n
        ('Python Files', '*.py'),\n
        ('Text Document', '*.txt')\n
    ]\n
    \n
    OverWrite:\n
    If True, file will be overwritten\n
    If False, only the path+filename will be returned\n
    """
    import tkinter as tk
    from tkinter.filedialog import asksaveasfile
    from tkinter.filedialog import askopenfilename

    # Create the window
    root = tk.Tk()
    # Hide the window
    root.withdraw()

    if OverWrite is True:
        file = asksaveasfile(filetypes=files, defaultextension=files)
        if file is not None:
            return file.name
    if OverWrite is False:
        file = askopenfilename(filetypes=files, defaultextension=files)
        if file is not None:
            return file


def GetLetterFromNumber(number: int, UCase: bool = True):
    # from openpyxl.utils import get_column_letter

    # Letter = get_column_letter(number)

    # # If UCase is true, convert to upper case
    # if UCase is True:
    #     Letter = Letter.upper()

    """Number to Excel-style column name, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA."""
    Letter = ""
    while number > 0:
        number, r = divmod(number - 1, 26)
        Letter = chr(r + ord("A")) + Letter
    return Letter


def GetNumberFromLetter(Letter):
    # from openpyxl.utils import column_index_from_string

    # Number = column_index_from_string(Letter)

    """Excel-style column name to number, e.g., A = 1, Z = 26, AA = 27, AAA = 703."""
    number = 0
    for c in Letter:
        number = number * 26 + 1 + ord(c) - ord("A")
    return number


def GetA1fromR1C1(input: str) -> str:
    if input[:1] == "'":
        input = input[1:]
    try:
        input = input.upper()
        ColumnPosition = input.find("C")
        RowNumber = int(input[1:(ColumnPosition)])
        ColumnNumber = int(input[(ColumnPosition + 1) :])

        ColumnLetter = GetLetterFromNumber(ColumnNumber)

        return str(ColumnLetter + str(RowNumber))
    except Exception:
        return ""


def RemoveNumbersFromString(string: str) -> str:
    no_digits = []

    # Iterate through the string, adding non-numbers to the no_digits list
    for i in string:
        if not i.isdigit():
            no_digits.append(i)

    # Now join all elements of the list with '',
    # which puts all of the characters together.
    result = "".join(no_digits)

    return result


def RemoveLettersFromString(string: str) -> str:
    no_chars = []

    # Iterate through the string, adding non-numbers to the no_digits list
    for i in string:
        if i.isdigit():
            no_chars.append(i)

    # Now join all elements of the list with '',
    # which puts all of the characters together.
    result = "".join(no_chars)

    return result


def CheckIfWorkbookExists(FullFileName: str, CreateIfNone: bool = True):
    import os

    try:
        from openpyxl import Workbook
    except Exception:
        return False

    result = False
    try:
        result = os.path.exists(FullFileName)
    except Exception:
        if CreateIfNone is True:
            Filter = [
                ("Excel", "*.xlsx"),
                (
                    "Excel Macro-enabled Workbook",
                    "*.xlsm",
                ),
            ]
            FullFileName = SaveDialog(Filter)
            if FullFileName.strip():
                wb = Workbook(str(FullFileName))
                wb.save(FullFileName)
                wb.close()
                result = True
        if CreateIfNone is False:
            result = False
    return result


def ColorConvertor(ColorRGB: [], Alpha: float = 1) -> ():
    """
    A single function to convert colors to rgba colors as a tuple of float from 0-1
    ColorRGB:   [255,255,255]
    Alpha:      0-1
    """
    from matplotlib import colors as mcolors

    ColorRed = ColorRGB[0] / 255
    colorGreen = ColorRGB[1] / 255
    colorBlue = ColorRGB[2] / 255

    color = (ColorRed, colorGreen, colorBlue)

    result = mcolors.to_rgba(color, Alpha)

    return result


def OpenFile(FileName: str):
    """
    Filename = full path with filename as string
    """
    import subprocess
    import os
    import platform

    try:
        if os.path.exists(FileName):
            if platform.system() == "Darwin":  # macOS
                subprocess.call(("open", FileName))
            elif platform.system() == "Windows":  # Windows
                os.startfile(FileName)
            else:  # linux variants
                print(FileName)
                try:
                    subprocess.check_output(["xdg-open", FileName.strip()])
                except subprocess.CalledProcessError:
                    Print(
                        f"An error occured when opening {FileName}!\n"
                        + "This can happen when running FreeCAD as an AppImage.\n"
                        + "Please install FreeCAD directly.",
                        "Error",
                    )
        else:
            print(f"Error: {FileName} does not exist.")
    except Exception as e:
        raise e


def SetColumnWidth_SpreadSheet(sheet, column: str, cellValue: str, factor: int = 10) -> bool:
    """_summary_

    Args:
        sheet (_type_): FreeCAD spreadsheet object.\n
        column (str): The column for which the width will be set. must be like "A", "B", etc.\n
        cellValue (str): The string to calulate the widht from.\n
        factor (int, optional): to increase the stringlength with a factor. Defaults to 10.\n

    Returns:
        bool: returns True or False
    """
    try:
        # Calculate the text length needed.
        length = int(len(cellValue) * factor)

        # Set the column width
        sheet.setColumnWidth(column, length)

        # Recompute the sheet
        sheet.recompute()
    except Exception:
        return False

    return True


def Print(Input: str, Type: str = ""):
    """_summary_

    Args:
        Input (str): Text to print.\n
        Type (str, optional): Type of message. (enter Warning, Error or Log). Defaults to "".
    """
    import FreeCAD as App

    if Type == "Warning":
        App.Console.PrintWarning(Input + "\n")
    elif Type == "Error":
        App.Console.PrintError(Input + "\n")
    elif Type == "Log":
        App.Console.PrintLog(Input + "\n")
    else:
        App.Console.PrintMessage(Input + "\n")


def LightOrDark(rgbColor=[0, 128, 255, 255]):
    """_summary_
    reference: https://alienryderflex.com/hsp.html
    Args:
        rgbColor (list, optional): RGB color. Defaults to [0, 128, 255, 255].\n
        note: The alpha value is added for completeness, but us ignored in the equation.

    Returns:
        string: "light or dark"
    """
    [r, g, b, a] = rgbColor
    hsp = math.sqrt(0.299 * (r * r) + 0.587 * (g * g) + 0.114 * (b * b))
    if hsp > 127.5:
        return "light"
    else:
        return "dark"


def toggleToolbars(ToolbarName: str, WorkBench: str = ""):
    import FreeCADGui as Gui
    from PySide.QtWidgets import QToolBar

    # Get the active workbench
    if WorkBench == "":
        WB = Gui.activeWorkbench()
    if WorkBench != "":
        WB = Gui.getWorkbench(WorkBench)

    # Get the list of toolbars present.
    ListToolbars = WB.listToolbars()
    # Go through the list. If the toolbar exists set ToolbarExists to True
    ToolbarExists = False
    for i in range(len(ListToolbars)):
        if ListToolbars[i] == ToolbarName:
            ToolbarExists = True

    # If ToolbarExists is True continue. Otherwise return.
    if ToolbarExists is True:
        # Get the main window
        mainWindow = Gui.getMainWindow()
        # Get the toolbar
        ToolBar = mainWindow.findChild(QToolBar, ToolbarName)
        # If the toolbar is not hidden, hide it and return.
        if ToolBar.isHidden() is False:
            ToolBar.setHidden(True)
            return
        # If the toolbar is hidden, set visible and return.
        if ToolBar.isHidden() is True:
            ToolBar.setVisible(True)
            return
    return


def PartFeatureList():
    result = [
        "Part::PartFeature",
        "Part::Extrusion",
        "Part::CustomFeature",
        "Part::Chamfer",
        "Part::Compound",
        "Part::Fillet",
        "Part::FeatureGeometrySet",
        "Part::Mirroring",
        "Part::Boolean",
        "Part::Box",
        "Part::Circle",
        "Part::MultiCommon",
        "Part::CurveNet",
        "Part::MultiFuse",
        "Part::ImportBrep",
        "Part::ImportIges",
        "Part::ImportStep",
        "Part::Polygon",
        "Part::RuledSurface",
        "Part::Loft",
        "Part::Sweep",
        "Part::Thickness",
        "Part::Refine",
        "Part::Reverse",
        "Part::Vertex",
        "Part::Line",
        "Part::Plane",
        "Part::Sphere",
        "Part::Ellipsoid",
        "Part::Cylinder",
        "Part::Prism",
        "Part::RegularPolygon",
        "Part::Cone",
        "Part::Torus",
        "Part::Helix",
        "Part::Spiral",
        "Part::Wedge",
        "Part::Ellipse",
        "Part::Extrusion",
        "Part::Face",
        "Part::Offset",
        "Part::Offset2D",
        "Part::Revolution",
        "Part::Part2DObject",
        "Part::Primitive",
    ]

    return result


def PartDesingFeatureList():
    result = [
        "PartDesign::Pad",
        "PartDesign::Pocket",
        "PartDesign::FeaturePrimitive",
        "PartDesign::FeatureBase",
        "PartDesign::Draft",
        "PartDesign::Fillet",
        "PartDesign::Groove",
        "PartDesign::Helix",
        "PartDesign::Hole",
        "PartDesign::Loft",
        "PartDesign::Pipe",
        "PartDesign::Box",
        "PartDesign::Cylinder",
        "PartDesign::Sphere",
        "PartDesign::Cone",
        "PartDesign::Ellipsoid",
        "PartDesign::Torus",
        "PartDesign::Prism",
        "PartDesign::Wedge",
        "PartDesign::Revolution",
        "PartDesign::Thickness",
        "PartDesign::Transformed",
        "PartDesign::Body",
        "PartDesign::Boolean",
        "PartDesign::Chamfer",
        "PartDesign::ShapeBinder",
        "PartDesign::SubShapeBinder",
    ]

    return result


def AllFeaturesList():
    result = [
        "PartDesign::Pad",
        "PartDesign::Pocket",
        "PartDesign::FeaturePrimitive",
        "Drawing::FeatureProjection",
        "Part::CustomFeature",
        "Part::Chamfer",
        "Part::Compound",
        "Part::Fillet",
        "Part::FeatureGeometrySet",
        "Part::Mirroring",
        "Part::Boolean",
        "Part::Box",
        "Part::Circle",
        "Part::MultiCommon",
        "Part::CurveNet",
        "Part::MultiFuse",
        "Part::ImportBrep",
        "Part::ImportIges",
        "Part::ImportStep",
        "Part::Polygon",
        "Part::RuledSurface",
        "Part::Loft",
        "Part::Sweep",
        "Part::Thickness",
        "Part::Refine",
        "Part::Reverse",
        "Part::Vertex",
        "Part::Line",
        "Part::Plane",
        "Part::Sphere",
        "Part::Ellipsoid",
        "Part::Cylinder",
        "Part::Prism",
        "Part::RegularPolygon",
        "Part::Cone",
        "Part::Torus",
        "Part::Helix",
        "Part::Spiral",
        "Part::Wedge",
        "Part::Ellipse",
        "PartDesign::FeatureBase",
        "PartDesign::Draft",
        "PartDesign::Fillet",
        "PartDesign::Groove",
        "PartDesign::Helix",
        "PartDesign::Hole",
        "PartDesign::Loft",
        "PartDesign::Pipe",
        "PartDesign::Box",
        "PartDesign::Cylinder",
        "PartDesign::Sphere",
        "PartDesign::Cone",
        "PartDesign::Ellipsoid",
        "PartDesign::Torus",
        "PartDesign::Prism",
        "PartDesign::Wedge",
        "PartDesign::Revolution",
        "PartDesign::Thickness",
        "PartDesign::Transformed",
        "Path::FeatureArea",
        "Path::FeatureAreaView",
        "Sketcher::SketchObjectSF",
        "Surface::Cut",
        "Surface::Filling",
        "Surface::GeomFillSurface",
        "Surface::Sewing",
        "TechDraw::FeatureProjection",
        "Part::Extrusion",
        "Part::Face",
        "Part::Offset",
        "Part::Offset2D",
        "Part::Revolution",
        "Part::Part2DObject",
        "Part::Primitive",
        "PartDesign::Body",
        "PartDesign::Boolean",
        "PartDesign::Chamfer",
        "PartDesign::ShapeBinder",
        "PartDesign::SubShapeBinder",
        "Sketcher::SketchObject",
        "Surface::Extend",
        "Surface::Sections.",
    ]

    return result
